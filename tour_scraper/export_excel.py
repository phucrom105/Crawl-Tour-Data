# export_excel.py
import json
import pandas as pd
from datetime import datetime

def export_to_excel(json_file='tours_data.json', excel_file='tours_data.xlsx'):
    """Chuyển đổi JSON sang Excel với format đẹp"""
    
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    processed_data = []
    for tour in data:
        # Format lịch trình - KHÔNG dùng \n làm separator
        lich_trinh_text = ''
        if tour.get('lich_trinh'):
            for day_idx, day in enumerate(tour['lich_trinh'], 1):
                # Thêm separator giữa các ngày
                if day_idx > 1:
                    lich_trinh_text += '\n' + '='*60 + '\n'
                
                # Thêm tiêu đề ngày
                lich_trinh_text += day.get('ngay', f'NGÀY {day_idx}') + '\n'
                lich_trinh_text += '='*60 + '\n'
                
                # Thêm các hoạt động - mỗi activity đã là một string hoàn chỉnh
                for activity in day.get('hoat_dong', []):
                    # Activity đã có format: "Sáng:\n* item1\n* item2"
                    # Chúng ta chỉ cần thêm vào và xuống dòng
                    lich_trinh_text += activity + '\n\n'
        
        # Làm phẳng các list khác
        def flatten_list(items, prefix='- '):
            if not items:
                return ''
            if isinstance(items, list):
                return '\n'.join([f"{prefix}{item}" for item in items])
            return str(items)
        
        processed_tour = {
            'URL': tour.get('url'),
            'Hình ảnh': tour.get('hinh_anh_chinh'),
            'Tour': tour.get('tour_name'), 
            'Tên Tour': tour.get('title'),
            'Mã Tour': tour.get('ma_tour'),
            'Thời gian': tour.get('thoi_gian'),
            'Khởi hành': tour.get('khoi_hanh'),
            'Vận chuyển': tour.get('van_chuyen'),
            'Xuất phát': tour.get('xuat_phat'),
            'Giá từ': tour.get('gia_tu'),
            'Trải nghiệm': flatten_list(tour.get('trai_nghiem')),
            'Điểm nhấn hành trình': flatten_list(tour.get('diem_nhan_hanh_trinh')),
            'Lịch trình': lich_trinh_text.strip(),
            'Dịch vụ bao gồm': flatten_list(tour.get('dich_vu_bao_gom')),
            'Dịch vụ không bao gồm': flatten_list(tour.get('dich_vu_khong_bao_gom')),
            'Ghi chú': flatten_list(tour.get('ghi_chu')),
        }
        processed_data.append(processed_tour)
    
    df = pd.DataFrame(processed_data)
    
    columns_order = [
       'URL', 'Hình ảnh', 'Tour', 'Tên Tour', 'Mã Tour', 
        'Thời gian', 'Khởi hành', 'Vận chuyển', 'Xuất phát', 'Giá từ',
        'Trải nghiệm', 'Điểm nhấn hành trình',
        'Lịch trình',
        'Dịch vụ bao gồm',
        'Dịch vụ không bao gồm',
        'Ghi chú'
    ]
    
    columns_order = [col for col in columns_order if col in df.columns]
    df = df[columns_order]
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Tours')
        
        worksheet = writer.sheets['Tours']
        
        # Set column widths - tùy chỉnh width cho từng cột
        column_widths = {
            'URL': 60,
            'Hình ảnh': 60,
            'Tour': 20,
            'Tên Tour': 60,
            'Mã Tour': 15,
            'Thời gian': 15,
            'Khởi hành': 25,
            'Vận chuyển': 25,
            'Xuất phát': 20,
            'Giá từ': 15,
            'Trải nghiệm': 50,
            'Điểm nhấn hành trình': 50,
            'Lịch trình': 80,
            'Dịch vụ bao gồm': 50,
            'Dịch vụ không bao gồm': 50,
            'Ghi chú': 50
        }
        
        for idx, col in enumerate(df.columns):
            # Lấy width tùy chỉnh hoặc tính toán tự động
            if col in column_widths:
                width = column_widths[col]
            else:
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                width = min(max_length + 2, 80)
            
            # Tính column letter
            if idx < 26:
                col_letter = chr(65 + idx)
            else:
                col_letter = chr(65 + idx // 26 - 1) + chr(65 + idx % 26)
            
            worksheet.column_dimensions[col_letter].width = width
        
        # Enable text wrap cho tất cả cells và style cho header
        from openpyxl.styles import Alignment, Font, PatternFill
        
        # Style cho header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
        
        # Style cho data cells
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
    
    print(f"✓ Đã xuất {len(processed_data)} tours ra file {excel_file}")
    print(f"✓ Các trường dữ liệu: {', '.join(columns_order)}")
    print(f"✓ File Excel đã được format với:")
    print(f"  - Header row màu xanh navy, chữ trắng, in đậm")
    print(f"  - Freeze header row")
    print(f"  - Wrap text tất cả cells")
    print(f"  - Column widths được tối ưu")

if __name__ == '__main__':
    export_to_excel()